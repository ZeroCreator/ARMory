import io
import re
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update
from sqlalchemy.orm import selectinload
from typing import List, Optional
from openpyxl import Workbook, load_workbook

from app.database import get_db
from app.models import GlossaryTerm, GlossaryTopic, GlossarySubtopic
from app.schemas import (
    GlossaryTermOut, GlossaryTermCreate, GlossaryTermUpdate,
    GlossaryTopicOut, GlossaryTopicCreate, GlossaryTopicUpdate,
    GlossarySubtopicOut, GlossarySubtopicCreate, GlossarySubtopicUpdate,
)


def _ilike(field, value: str):
    """Регистронезависимое вхождение, совместимое с Unicode в SQLite."""
    return func.lower(field).like(f"%{value.lower()}%")


router = APIRouter(prefix="/api/glossary", tags=["glossary"])


def _normalize_letter(term: str) -> str:
    if not term:
        return "#"
    first = term.strip()[0].upper()
    first = first.replace("Ё", "Е")
    if re.match(r"[А-ЯA-Z]", first):
        return first
    return "#"


def _natural_sort_key(name: str):
    """Ключ для естественной сортировки строк с цифрами (1, 2, 10, а не 1, 10, 2)."""
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", name or "")
    ]


async def _set_term_topic_subtopic(term: GlossaryTerm, data: GlossaryTermUpdate, db: AsyncSession):
    """Приводит topic_id/subtopic_id в согласованное состояние."""
    topic_id = data.topic_id if data.topic_id is not None else term.topic_id

    if data.subtopic_id is not None:
        if data.subtopic_id == 0:
            term.subtopic_id = None
        else:
            sub = await db.get(GlossarySubtopic, data.subtopic_id)
            if not sub:
                raise HTTPException(status_code=404, detail="Subtopic not found")
            term.subtopic_id = sub.id
            topic_id = sub.topic_id

    # Если тема изменилась и не совпадает с текущей подтемой — сбрасываем подтему
    if topic_id != term.topic_id and term.subtopic_id:
        sub = await db.get(GlossarySubtopic, term.subtopic_id)
        if not sub or sub.topic_id != topic_id:
            term.subtopic_id = None

    term.topic_id = topic_id if topic_id else None


# ═══════════════════════════════════════════════════
# Темы глоссария
# ═══════════════════════════════════════════════════

@router.get("/topics", response_model=List[GlossaryTopicOut])
async def list_glossary_topics(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(GlossaryTopic)
        .options(selectinload(GlossaryTopic.subtopics))
        .order_by(GlossaryTopic.sort_order)
    )
    topics = sorted(result.scalars().all(), key=lambda t: (t.sort_order, _natural_sort_key(t.name)))

    topic_counts = {}
    subtopic_counts = {}
    counts_result = await db.execute(
        select(GlossaryTerm.topic_id, GlossaryTerm.subtopic_id, func.count(GlossaryTerm.id))
        .group_by(GlossaryTerm.topic_id, GlossaryTerm.subtopic_id)
    )
    for row in counts_result.all():
        t_id, s_id, cnt = row
        if t_id is not None:
            topic_counts[t_id] = topic_counts.get(t_id, 0) + cnt
        if s_id is not None:
            subtopic_counts[s_id] = subtopic_counts.get(s_id, 0) + cnt

    return [
        GlossaryTopicOut(
            id=t.id,
            name=t.name,
            sort_order=t.sort_order,
            term_count=topic_counts.get(t.id, 0),
            created_at=t.created_at,
            subtopics=[
                GlossarySubtopicOut(
                    id=s.id,
                    topic_id=s.topic_id,
                    name=s.name,
                    sort_order=s.sort_order,
                    term_count=subtopic_counts.get(s.id, 0),
                    created_at=s.created_at,
                )
                for s in sorted(t.subtopics, key=lambda x: (x.sort_order, _natural_sort_key(x.name)))
            ],
        )
        for t in topics
    ]


@router.post("/topics", response_model=GlossaryTopicOut, status_code=201)
async def create_glossary_topic(data: GlossaryTopicCreate, db: AsyncSession = Depends(get_db)):
    existing = await db.execute(select(GlossaryTopic).where(func.lower(GlossaryTopic.name) == data.name.lower()))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Topic already exists")
    topic = GlossaryTopic(name=data.name.strip())
    db.add(topic)
    await db.commit()
    await db.refresh(topic)
    return GlossaryTopicOut.model_validate(topic)


@router.patch("/topics/{topic_id}", response_model=GlossaryTopicOut)
async def update_glossary_topic(
    topic_id: int,
    data: GlossaryTopicUpdate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(GlossaryTopic).where(GlossaryTopic.id == topic_id))
    topic = result.scalar_one_or_none()
    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")

    if data.name is not None:
        name = data.name.strip()
        existing = await db.execute(
            select(GlossaryTopic)
            .where(func.lower(GlossaryTopic.name) == name.lower())
            .where(GlossaryTopic.id != topic_id)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Topic already exists")
        topic.name = name
    if data.sort_order is not None:
        topic.sort_order = data.sort_order

    await db.commit()
    await db.refresh(topic)
    return GlossaryTopicOut.model_validate(topic)


@router.delete("/topics/{topic_id}", status_code=204)
async def delete_glossary_topic(topic_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(GlossaryTopic).where(GlossaryTopic.id == topic_id))
    topic = result.scalar_one_or_none()
    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")

    # Перед удалением темы сбрасываем term-ы в "без темы"
    await db.execute(
        update(GlossaryTerm)
        .where(GlossaryTerm.topic_id == topic_id)
        .values(topic_id=None, subtopic_id=None)
    )
    await db.commit()
    await db.delete(topic)
    await db.commit()
    return None


# ═══════════════════════════════════════════════════
# Подтемы глоссария
# ═══════════════════════════════════════════════════

@router.get("/subtopics", response_model=List[GlossarySubtopicOut])
async def list_glossary_subtopics(topic_id: Optional[int] = None, db: AsyncSession = Depends(get_db)):
    query = select(GlossarySubtopic)
    if topic_id is not None:
        query = query.where(GlossarySubtopic.topic_id == topic_id)
    query = query.order_by(GlossarySubtopic.sort_order)
    result = await db.execute(query)
    subtopics = sorted(result.scalars().all(), key=lambda s: (s.sort_order, _natural_sort_key(s.name)))

    counts_result = await db.execute(
        select(GlossaryTerm.subtopic_id, func.count(GlossaryTerm.id))
        .group_by(GlossaryTerm.subtopic_id)
    )
    counts = {s_id: cnt for s_id, cnt in counts_result.all() if s_id is not None}

    return [
        GlossarySubtopicOut(
            id=s.id,
            topic_id=s.topic_id,
            name=s.name,
            sort_order=s.sort_order,
            term_count=counts.get(s.id, 0),
            created_at=s.created_at,
        )
        for s in subtopics
    ]


@router.post("/subtopics", response_model=GlossarySubtopicOut, status_code=201)
async def create_glossary_subtopic(data: GlossarySubtopicCreate, db: AsyncSession = Depends(get_db)):
    topic = await db.get(GlossaryTopic, data.topic_id)
    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")

    existing = await db.execute(
        select(GlossarySubtopic)
        .where(GlossarySubtopic.topic_id == data.topic_id)
        .where(func.lower(GlossarySubtopic.name) == data.name.lower())
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Subtopic already exists in this topic")

    subtopic = GlossarySubtopic(topic_id=data.topic_id, name=data.name.strip())
    db.add(subtopic)
    await db.commit()
    await db.refresh(subtopic)
    return GlossarySubtopicOut.model_validate(subtopic)


@router.patch("/subtopics/{subtopic_id}", response_model=GlossarySubtopicOut)
async def update_glossary_subtopic(
    subtopic_id: int,
    data: GlossarySubtopicUpdate,
    db: AsyncSession = Depends(get_db),
):
    subtopic = await db.get(GlossarySubtopic, subtopic_id)
    if not subtopic:
        raise HTTPException(status_code=404, detail="Subtopic not found")

    target_topic_id = data.topic_id if data.topic_id is not None else subtopic.topic_id

    if data.name is not None:
        name = data.name.strip()
        existing = await db.execute(
            select(GlossarySubtopic)
            .where(GlossarySubtopic.topic_id == target_topic_id)
            .where(func.lower(GlossarySubtopic.name) == name.lower())
            .where(GlossarySubtopic.id != subtopic_id)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Subtopic already exists in this topic")
        subtopic.name = name

    if data.topic_id is not None:
        topic = await db.get(GlossaryTopic, data.topic_id)
        if not topic:
            raise HTTPException(status_code=404, detail="Topic not found")
        subtopic.topic_id = data.topic_id

    if data.sort_order is not None:
        subtopic.sort_order = data.sort_order

    await db.commit()
    await db.refresh(subtopic)
    return GlossarySubtopicOut.model_validate(subtopic)


@router.delete("/subtopics/{subtopic_id}", status_code=204)
async def delete_glossary_subtopic(subtopic_id: int, db: AsyncSession = Depends(get_db)):
    subtopic = await db.get(GlossarySubtopic, subtopic_id)
    if not subtopic:
        raise HTTPException(status_code=404, detail="Subtopic not found")

    await db.execute(
        update(GlossaryTerm)
        .where(GlossaryTerm.subtopic_id == subtopic_id)
        .values(subtopic_id=None)
    )
    await db.commit()
    await db.delete(subtopic)
    await db.commit()
    return None


# ═══════════════════════════════════════════════════
# Термины глоссария
# ═══════════════════════════════════════════════════

@router.get("", response_model=List[GlossaryTermOut])
async def list_glossary_terms(
    q: Optional[str] = None,
    letter: Optional[str] = None,
    topic_id: Optional[int] = None,
    subtopic_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
):
    query = select(GlossaryTerm).options(
        selectinload(GlossaryTerm.topic),
        selectinload(GlossaryTerm.subtopic),
    )
    if q:
        query = query.where(
            _ilike(GlossaryTerm.term, q) | _ilike(GlossaryTerm.short_definition, q) | _ilike(GlossaryTerm.definition, q)
        )
    if letter:
        query = query.where(GlossaryTerm.letter == letter.upper())
    if topic_id is not None:
        if topic_id == 0:
            query = query.where(GlossaryTerm.topic_id.is_(None))
        else:
            query = query.where(GlossaryTerm.topic_id == topic_id)
    if subtopic_id is not None:
        if subtopic_id == 0:
            query = query.where(GlossaryTerm.subtopic_id.is_(None))
        else:
            query = query.where(GlossaryTerm.subtopic_id == subtopic_id)
    query = query.order_by(GlossaryTerm.letter, GlossaryTerm.term, GlossaryTerm.sort_order)
    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/count", response_model=int)
async def count_glossary_terms(
    q: Optional[str] = None,
    topic_id: Optional[int] = None,
    subtopic_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    query = select(func.count(GlossaryTerm.id))
    if q:
        query = query.where(
            _ilike(GlossaryTerm.term, q) | _ilike(GlossaryTerm.short_definition, q) | _ilike(GlossaryTerm.definition, q)
        )
    if topic_id is not None:
        if topic_id == 0:
            query = query.where(GlossaryTerm.topic_id.is_(None))
        else:
            query = query.where(GlossaryTerm.topic_id == topic_id)
    if subtopic_id is not None:
        if subtopic_id == 0:
            query = query.where(GlossaryTerm.subtopic_id.is_(None))
        else:
            query = query.where(GlossaryTerm.subtopic_id == subtopic_id)
    result = await db.execute(query)
    return result.scalar_one()


@router.get("/export")
async def export_glossary(db: AsyncSession = Depends(get_db)):
    """Экспорт глоссария в Excel (.xlsx)."""
    wb = Workbook()
    wb.remove(wb.active)

    # --- Термины (сначала, чтобы открывалась эта вкладка) ---
    ws_terms = wb.create_sheet("terms")
    ws_terms.append(["id", "term", "short_definition", "definition", "topic_name", "subtopic_name", "sort_order"])
    result = await db.execute(
        select(GlossaryTerm).options(selectinload(GlossaryTerm.topic), selectinload(GlossaryTerm.subtopic))
        .order_by(GlossaryTerm.letter, GlossaryTerm.term)
    )
    terms = result.scalars().all()
    for term in terms:
        ws_terms.append([
            term.id,
            term.term,
            term.short_definition or "",
            term.definition or "",
            term.topic.name if term.topic else "",
            term.subtopic.name if term.subtopic else "",
            term.sort_order,
        ])

    # --- Темы ---
    ws_topics = wb.create_sheet("topics")
    ws_topics.append(["id", "name", "sort_order"])
    result = await db.execute(select(GlossaryTopic).order_by(GlossaryTopic.sort_order))
    topics = result.scalars().all()
    topic_name_by_id = {}
    for t in topics:
        ws_topics.append([t.id, t.name, t.sort_order])
        topic_name_by_id[t.id] = t.name

    # --- Подтемы ---
    ws_subtopics = wb.create_sheet("subtopics")
    ws_subtopics.append(["id", "topic_name", "name", "sort_order"])
    result = await db.execute(
        select(GlossarySubtopic).options(selectinload(GlossarySubtopic.topic)).order_by(GlossarySubtopic.sort_order)
    )
    subtopics = result.scalars().all()
    for s in subtopics:
        ws_subtopics.append([s.id, s.topic.name if s.topic else "", s.name, s.sort_order])

    wb.active = ws_terms

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=glossary_export.xlsx"},
    )


@router.post("/import")
async def import_glossary(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Импорт глоссария из Excel (.xlsx)."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {e}")

    # Загружаем существующие данные для дедупликации по имени
    existing_topics = {
        t.name.lower(): t for t in (await db.execute(select(GlossaryTopic))).scalars().all()
    }
    existing_subtopics = {
        ((s.topic.name.lower() if s.topic else ""), s.name.lower()): s
        for s in (await db.execute(select(GlossarySubtopic).options(selectinload(GlossarySubtopic.topic)))).scalars().all()
    }
    existing_terms = {
        t.term.lower(): t for t in (await db.execute(select(GlossaryTerm))).scalars().all()
    }

    topic_name_to_id: dict[str, int] = {}
    subtopic_key_to_id: dict[tuple[str, str], int] = {}
    created_topics = created_subtopics = created_terms = updated_terms = 0

    # --- Темы ---
    ws_topics = wb["topics"] if "topics" in wb.sheetnames else None
    if ws_topics:
        for row in ws_topics.iter_rows(min_row=2, values_only=True):
            name = str(row[1] or "").strip()
            if not name:
                continue
            sort_order = row[2] if row[2] is not None else 0
            key = name.lower()
            if key in existing_topics:
                topic_name_to_id[name] = existing_topics[key].id
            else:
                topic = GlossaryTopic(name=name, sort_order=sort_order)
                db.add(topic)
                await db.flush()
                topic_name_to_id[name] = topic.id
                existing_topics[key] = topic
                created_topics += 1

    # --- Подтемы ---
    ws_subtopics = wb["subtopics"] if "subtopics" in wb.sheetnames else None
    if ws_subtopics:
        for row in ws_subtopics.iter_rows(min_row=2, values_only=True):
            topic_name = str(row[1] or "").strip()
            name = str(row[2] or "").strip()
            if not name:
                continue
            sort_order = row[3] if row[3] is not None else 0
            topic_id = topic_name_to_id.get(topic_name)
            if topic_id is None:
                continue
            key = (topic_name.lower(), name.lower())
            if key in existing_subtopics:
                subtopic_key_to_id[key] = existing_subtopics[key].id
            else:
                subtopic = GlossarySubtopic(topic_id=topic_id, name=name, sort_order=sort_order)
                db.add(subtopic)
                await db.flush()
                subtopic_key_to_id[key] = subtopic.id
                existing_subtopics[key] = subtopic
                created_subtopics += 1

    # --- Термины ---
    ws_terms = wb["terms"] if "terms" in wb.sheetnames else None
    if ws_terms:
        for row in ws_terms.iter_rows(min_row=2, values_only=True):
            term_text = str(row[1] or "").strip()
            if not term_text:
                continue
            short_definition = str(row[2] or "").strip()
            definition = str(row[3] or "").strip()
            topic_name = str(row[4] or "").strip()
            subtopic_name = str(row[5] or "").strip()
            sort_order = row[6] if row[6] is not None else 0

            topic_id = topic_name_to_id.get(topic_name)
            subtopic_id = None
            if subtopic_name:
                subtopic_id = subtopic_key_to_id.get((topic_name.lower(), subtopic_name.lower()))

            key = term_text.lower()
            if key in existing_terms:
                term = existing_terms[key]
                term.short_definition = short_definition
                term.definition = definition
                term.topic_id = topic_id
                term.subtopic_id = subtopic_id
                term.sort_order = sort_order
                updated_terms += 1
            else:
                term = GlossaryTerm(
                    term=term_text,
                    short_definition=short_definition,
                    definition=definition,
                    letter=_normalize_letter(term_text),
                    topic_id=topic_id,
                    subtopic_id=subtopic_id,
                    sort_order=sort_order,
                )
                db.add(term)
                created_terms += 1

    await db.commit()
    return {
        "created_topics": created_topics,
        "created_subtopics": created_subtopics,
        "created_terms": created_terms,
        "updated_terms": updated_terms,
    }


@router.post("", response_model=GlossaryTermOut, status_code=201)
async def create_glossary_term(data: GlossaryTermCreate, db: AsyncSession = Depends(get_db)):
    term = GlossaryTerm(
        term=data.term,
        short_definition=data.short_definition,
        definition=data.definition,
        letter=_normalize_letter(data.term),
    )
    await _set_term_topic_subtopic(term, data, db)
    db.add(term)
    await db.commit()
    await db.refresh(term)
    result = await db.execute(
        select(GlossaryTerm)
        .options(selectinload(GlossaryTerm.topic), selectinload(GlossaryTerm.subtopic))
        .where(GlossaryTerm.id == term.id)
    )
    return result.scalar_one()


@router.get("/{term_id}", response_model=GlossaryTermOut)
async def get_glossary_term(term_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(GlossaryTerm)
        .options(selectinload(GlossaryTerm.topic), selectinload(GlossaryTerm.subtopic))
        .where(GlossaryTerm.id == term_id)
    )
    term = result.scalar_one_or_none()
    if not term:
        raise HTTPException(status_code=404, detail="Term not found")
    return term


@router.patch("/{term_id}", response_model=GlossaryTermOut)
async def update_glossary_term(
    term_id: int,
    data: GlossaryTermUpdate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(GlossaryTerm).where(GlossaryTerm.id == term_id))
    term = result.scalar_one_or_none()
    if not term:
        raise HTTPException(status_code=404, detail="Term not found")

    if data.term is not None:
        term.term = data.term
        term.letter = _normalize_letter(data.term)
    if data.short_definition is not None:
        term.short_definition = data.short_definition
    if data.definition is not None:
        term.definition = data.definition

    await _set_term_topic_subtopic(term, data, db)

    await db.commit()
    await db.refresh(term)
    result = await db.execute(
        select(GlossaryTerm)
        .options(selectinload(GlossaryTerm.topic), selectinload(GlossaryTerm.subtopic))
        .where(GlossaryTerm.id == term.id)
    )
    return result.scalar_one()


@router.delete("/{term_id}", status_code=204)
async def delete_glossary_term(term_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(GlossaryTerm).where(GlossaryTerm.id == term_id))
    term = result.scalar_one_or_none()
    if not term:
        raise HTTPException(status_code=404, detail="Term not found")
    await db.delete(term)
    await db.commit()
    return None
