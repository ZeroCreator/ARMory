from datetime import datetime
from typing import Optional

from pydantic import BaseModel

import io
import os
import platform
import subprocess
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles.colors import Color
from sqlalchemy import select, update, delete, func, distinct, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import Settings, get_settings
from app.database import get_db
from app.events import broadcast
from app.models import Assignee, Project, Task, TaskAssignee, TaskAttachment, TaskStatus, TaskStatusHistory
from app.routers.collabora import build_collabora_iframe_url
from app.routers.wopi import OFFICE_EXTENSIONS, encode_file_id
from app.storage import StorageBackend, get_storage
from app.schemas import (
    GlobalKanbanColumnCreate,
    GlobalKanbanColumnUpdate,
    KanbanBoardOut,
    KanbanColumnOut,
    KanbanExportOut,
    KanbanFiltersOut,
    KanbanGlobalOut,
    KanbanImportIn,
    KanbanProjectExport,
    KanbanTaskExport,
    KanbanTaskStatusUpdate,
    TaskAttachmentCreate,
    TaskAttachmentOut,
    TaskAttachmentUpdate,
    TaskBulkAttachment,
    TaskBulkAddAttachmentsRequest,
    TaskBulkCreate,
    TaskBulkOut,
    TaskBulkRequest,
    TaskBulkUpdate,
    TaskBulkUpdateRequest,
    TaskCreate,
    TaskOut,
    TaskReorderRequest,
    TaskStatusCreate,
    TaskStatusHistoryOut,
    TaskStatusOut,
    TaskStatusReorderRequest,
    TaskStatusUpdate,
    TaskUpdate,
    TaskListTelegramConfig,
)

router = APIRouter(prefix="/api/projects/{project_id}", tags=["tasks"])
global_router = APIRouter(prefix="/api", tags=["kanban"])


class CurrentUserOut(BaseModel):
    email: Optional[str] = None


@global_router.get("/me", response_model=CurrentUserOut)
async def get_current_user(request: Request):
    """Возвращает email текущего пользователя из заголовков oauth2-proxy."""
    headers = [
        "X-Forwarded-Email",
        "X-Forwarded-User",
        "X-Forwarded-Preferred-Username",
        "X-Forwarded-Access-Token",
        "Remote-User",
        "Remote-Email",
    ]
    email = None
    for h in headers:
        value = request.headers.get(h)
        if value:
            email = value
            break
    return {"email": email.strip() if email else None}


@global_router.get("/me/debug")
async def debug_current_user(request: Request):
    """Отладочный endpoint: возвращает все потенциальные auth-заголовки."""
    auth_headers = {}
    for key, value in request.headers.items():
        if any(
            x in key.lower()
            for x in ["forwarded", "remote", "auth", "user", "email", "preferred"]
        ):
            auth_headers[key] = value
    return {"headers": auth_headers, "all": dict(request.headers)}


async def _get_project(project_id: int, db: AsyncSession) -> Project:
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


async def _get_status(project_id: int, status_id: int, db: AsyncSession) -> TaskStatus:
    result = await db.execute(
        select(TaskStatus).where(TaskStatus.id == status_id, TaskStatus.project_id == project_id)
    )
    status = result.scalar_one_or_none()
    if not status:
        raise HTTPException(status_code=404, detail="Task status not found")
    return status


async def _get_task(project_id: int, task_id: int, db: AsyncSession) -> Task:
    result = await db.execute(
        select(Task)
        .options(selectinload(Task.status), selectinload(Task.attachments), selectinload(Task.assignees))
        .where(Task.id == task_id, Task.project_id == project_id)
    )
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


def _extract_assignee_emails(data) -> list[str]:
    """Извлекает список email исполнителей из данных задачи."""
    emails = []
    if getattr(data, "assignee_emails", None):
        emails = [e.strip() for e in data.assignee_emails if e and e.strip()]
    if not emails and getattr(data, "assignee_email", None):
        emails = [data.assignee_email.strip()]
    return emails


async def _apply_task_assignees(task: Task, emails: list[str], db: AsyncSession) -> None:
    """Обновляет исполнителей задачи и legacy-поле assignee_email."""
    normalized = sorted(set(e.lower() for e in emails if e))
    task.assignee_email = normalized[0] if normalized else None

    existing = {a.assignee_email.lower(): a for a in task.assignees}
    to_keep = set()
    for email in normalized:
        if email in existing:
            to_keep.add(email)
        else:
            ta = TaskAssignee(task_id=task.id, assignee_email=email)
            db.add(ta)
            to_keep.add(email)

    for email, ta in existing.items():
        if email not in to_keep:
            await db.delete(ta)


async def _record_status_history(
    db: AsyncSession, task_id: int, status_id: int, entered_at: Optional[datetime] = None
) -> None:
    """Создать запись о переходе задачи в указанную колонку."""
    db.add(TaskStatusHistory(
        task_id=task_id,
        status_id=status_id,
        entered_at=entered_at or datetime.utcnow(),
    ))


def _task_has_assignee(task: Task, email: str) -> bool:
    """Проверяет, назначен ли указанный email исполнителем задачи."""
    if not email:
        return False
    email_lower = email.lower()
    if task.assignee_email and task.assignee_email.lower() == email_lower:
        return True
    return any(a.assignee_email.lower() == email_lower for a in task.assignees)


def _format_task_assignees(task: Task, assignees_map: dict[str, str]) -> str:
    """Возвращает строку с именами/емейлами исполнителей задачи через запятую."""
    names = []
    for email in task.assignee_emails:
        name = assignees_map.get(email) or email
        if name:
            names.append(name)
    return ", ".join(names) or "—"


# ═══════════════════════════════════════════════════
# Колонки (статусы)
# ═══════════════════════════════════════════════════

@router.get("/task-statuses", response_model=list[TaskStatusOut])
async def list_task_statuses(project_id: int, db: AsyncSession = Depends(get_db)):
    await _get_project(project_id, db)
    result = await db.execute(
        select(TaskStatus)
        .where(TaskStatus.project_id == project_id)
        .order_by(TaskStatus.sort_order.asc(), TaskStatus.created_at.asc())
    )
    return result.scalars().all()


@router.post("/task-statuses", response_model=TaskStatusOut, status_code=201)
async def create_task_status(
    project_id: int,
    data: TaskStatusCreate,
    db: AsyncSession = Depends(get_db),
):
    await _get_project(project_id, db)
    max_order = await db.execute(
        select(TaskStatus.sort_order)
        .where(TaskStatus.project_id == project_id)
        .order_by(TaskStatus.sort_order.desc())
        .limit(1)
    )
    max_val = max_order.scalar_one_or_none() or 0
    status = TaskStatus(
        project_id=project_id,
        name=data.name,
        color=data.color or "#a78bfa",
        sort_order=max_val + 1,
    )
    db.add(status)
    await db.commit()
    await db.refresh(status)
    broadcast({"type": "board_changed", "project_id": project_id})
    return status


@router.patch("/task-statuses/reorder", status_code=204)
async def reorder_task_statuses(
    project_id: int,
    data: TaskStatusReorderRequest,
    db: AsyncSession = Depends(get_db),
):
    await _get_project(project_id, db)
    for idx, status_id in enumerate(data.status_ids):
        await db.execute(
            update(TaskStatus)
            .where(TaskStatus.id == status_id, TaskStatus.project_id == project_id)
            .values(sort_order=idx)
        )
    await db.commit()
    broadcast({"type": "board_changed", "project_id": project_id})
    return None


@router.patch("/task-statuses/{status_id}", response_model=TaskStatusOut)
async def update_task_status(
    project_id: int,
    status_id: int,
    data: TaskStatusUpdate,
    db: AsyncSession = Depends(get_db),
):
    status = await _get_status(project_id, status_id, db)
    if data.name is not None:
        status.name = data.name
    if data.color is not None:
        status.color = data.color
    status.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(status)
    broadcast({"type": "board_changed", "project_id": project_id})
    return status


@router.delete("/task-statuses/{status_id}", status_code=204)
async def delete_task_status(
    project_id: int,
    status_id: int,
    db: AsyncSession = Depends(get_db),
):
    status = await _get_status(project_id, status_id, db)
    await db.delete(status)
    await db.commit()
    broadcast({"type": "board_changed", "project_id": project_id})
    return None


# ═══════════════════════════════════════════════════
# Задачи
# ═══════════════════════════════════════════════════

@router.get("/tasks", response_model=list[TaskOut])
async def list_tasks(project_id: int, db: AsyncSession = Depends(get_db)):
    await _get_project(project_id, db)
    result = await db.execute(
        select(Task)
        .options(selectinload(Task.status), selectinload(Task.attachments))
        .where(Task.project_id == project_id)
        .order_by(Task.is_closed.asc(), Task.sort_order.asc(), Task.created_at.asc())
    )
    return result.scalars().all()


@router.get("/tasks/board", response_model=KanbanBoardOut)
async def get_kanban_board(
    project_id: int,
    priority: Optional[str] = None,
    assignee_email: Optional[str] = None,
    tags: Optional[str] = None,
    created_after: Optional[datetime] = None,
    created_before: Optional[datetime] = None,
    due_after: Optional[datetime] = None,
    due_before: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
):
    await _get_project(project_id, db)
    statuses_result = await db.execute(
        select(TaskStatus)
        .where(TaskStatus.project_id == project_id)
        .order_by(TaskStatus.sort_order.asc(), TaskStatus.created_at.asc())
    )
    statuses = statuses_result.scalars().all()

    tasks_query = (
        select(Task)
        .options(selectinload(Task.status), selectinload(Task.attachments))
        .where(Task.project_id == project_id)
    )
    if priority is not None:
        tasks_query = tasks_query.where(Task.priority == priority)
    if assignee_email is not None:
        search = f"%{assignee_email}%"
        tasks_query = tasks_query.outerjoin(
            TaskAssignee, Task.id == TaskAssignee.task_id
        ).where(
            or_(
                Task.assignee_email.ilike(search),
                TaskAssignee.assignee_email.ilike(search),
            )
        ).distinct()
    if tags is not None:
        tasks_query = tasks_query.where(Task.tags.ilike(f"%{tags}%"))
    if created_after is not None:
        tasks_query = tasks_query.where(Task.created_at >= created_after)
    if created_before is not None:
        tasks_query = tasks_query.where(Task.created_at <= created_before)
    if due_after is not None:
        tasks_query = tasks_query.where(Task.due_date >= due_after)
    if due_before is not None:
        tasks_query = tasks_query.where(Task.due_date <= due_before)

    tasks_result = await db.execute(
        tasks_query.order_by(Task.is_closed.asc(), Task.sort_order.asc(), Task.created_at.asc())
    )
    tasks = tasks_result.scalars().all()

    return KanbanBoardOut(statuses=statuses, tasks=tasks)


@router.get("/kanban/filters", response_model=KanbanFiltersOut)
async def project_kanban_filters(project_id: int, db: AsyncSession = Depends(get_db)):
    """Доступные значения фильтров для kanban проекта."""
    await _get_project(project_id, db)

    priorities_result = await db.execute(
        select(distinct(Task.priority))
        .where(Task.priority.isnot(None), Task.project_id == project_id)
    )
    priorities = [p[0] for p in priorities_result.fetchall() if p[0]]

    assignees_result = await db.execute(select(Assignee).order_by(Assignee.name.asc()))
    assignees = assignees_result.scalars().all()

    tags_result = await db.execute(
        select(Task.tags).where(Task.tags.isnot(None), Task.project_id == project_id)
    )
    tag_set = set()
    for (tags_str,) in tags_result.fetchall():
        for tag in tags_str.split(","):
            tag = tag.strip()
            if tag:
                tag_set.add(tag)

    list_names_result = await db.execute(
        select(distinct(Task.list_name))
        .where(Task.list_name.isnot(None), Task.project_id == project_id)
    )
    list_names = [ln[0] for ln in list_names_result.fetchall() if ln[0]]

    return KanbanFiltersOut(
        projects=[],
        priorities=sorted(priorities),
        assignees=assignees,
        tags=sorted(tag_set),
        list_names=sorted(list_names),
    )


@router.post("/tasks", response_model=TaskOut, status_code=201)
async def create_task(
    project_id: int,
    data: TaskCreate,
    db: AsyncSession = Depends(get_db),
):
    await _get_project(project_id, db)
    await _get_status(project_id, data.status_id, db)

    max_order = await db.execute(
        select(Task.sort_order)
        .where(Task.project_id == project_id, Task.status_id == data.status_id)
        .order_by(Task.sort_order.desc())
        .limit(1)
    )
    max_val = max_order.scalar_one_or_none() or 0

    task = Task(
        project_id=project_id,
        status_id=data.status_id,
        title=data.title or "",
        description=data.description,
        priority=data.priority or "medium",
        is_closed=data.is_closed or False,
        start_date=data.start_date,
        due_date=None if data.is_closed else data.due_date,
        assignee_email=None,
        tags=data.tags,
        list_name=data.list_name,
        result=data.result,
        sort_order=max_val + 1,
    )
    db.add(task)
    await db.flush()
    await db.refresh(task)

    await _record_status_history(db, task.id, task.status_id, entered_at=task.created_at)

    assignee_emails = _extract_assignee_emails(data)
    await _apply_task_assignees(task, assignee_emails, db)

    await db.commit()
    await db.refresh(task)
    broadcast({"type": "task_changed", "project_id": project_id, "task_id": task.id, "status_id": task.status_id})
    return task


async def _get_or_create_todo_status(project_id: int, db: AsyncSession) -> TaskStatus:
    """Найти или создать статус 'К выполнению' в проекте."""
    result = await db.execute(
        select(TaskStatus)
        .where(TaskStatus.project_id == project_id, TaskStatus.name == "К выполнению")
    )
    status = result.scalar_one_or_none()
    if status:
        return status

    max_order = await db.execute(
        select(TaskStatus.sort_order)
        .where(TaskStatus.project_id == project_id)
        .order_by(TaskStatus.sort_order.desc())
        .limit(1)
    )
    max_val = max_order.scalar_one_or_none() or 0
    status = TaskStatus(
        project_id=project_id,
        name="К выполнению",
        color="#a78bfa",
        sort_order=max_val + 1,
    )
    db.add(status)
    await db.flush()
    await db.refresh(status)
    return status


async def _bulk_create_tasks(
    project_id: int,
    tasks_data: list[TaskBulkCreate],
    attachments_data: list[TaskBulkAttachment],
    db: AsyncSession,
) -> list[Task]:
    """Создать несколько задач. Если статус не указан — используется 'К выполнению'."""
    default_status = await _get_or_create_todo_status(project_id, db)

    # Определяем status_id для каждой задачи и считаем max_order внутри каждого статуса
    status_ids = []
    for task_data in tasks_data:
        status_id = task_data.status_id
        if status_id:
            await _get_status(project_id, status_id, db)
        else:
            status_id = default_status.id
        status_ids.append(status_id)

    max_orders: dict[int, int] = {}
    for status_id in set(status_ids):
        max_order_result = await db.execute(
            select(Task.sort_order)
            .where(Task.project_id == project_id, Task.status_id == status_id)
            .order_by(Task.sort_order.desc())
            .limit(1)
        )
        max_orders[status_id] = max_order_result.scalar_one_or_none() or 0

    created_tasks: list[Task] = []
    for idx, task_data in enumerate(tasks_data):
        status_id = status_ids[idx]
        max_orders[status_id] += 1
        task = Task(
            project_id=project_id,
            status_id=status_id,
            title=task_data.title or "",
            description=task_data.description,
            priority=task_data.priority or "medium",
            is_closed=False,
            start_date=task_data.start_date,
            due_date=task_data.due_date,
            assignee_email=None,
            tags=task_data.tags,
            list_name=task_data.list_name,
            sort_order=max_orders[status_id],
        )
        db.add(task)
        created_tasks.append(task)

    await db.flush()

    for task in created_tasks:
        await _record_status_history(db, task.id, task.status_id, entered_at=task.created_at)

    for task in created_tasks:
        assignee_emails = _extract_assignee_emails(task_data)
        await _apply_task_assignees(task, assignee_emails, db)
        for attachment_data in attachments_data:
            attachment = TaskAttachment(
                task_id=task.id,
                attachment_type=attachment_data.attachment_type,
                title=attachment_data.title,
                url=attachment_data.url,
                file_path=attachment_data.file_path,
            )
            db.add(attachment)

    await db.commit()

    for task in created_tasks:
        await db.refresh(task, attribute_names=["status", "attachments"])

    return created_tasks


@router.post("/tasks/bulk", response_model=TaskBulkOut, status_code=201)
async def create_tasks_bulk(
    project_id: int,
    data: TaskBulkRequest,
    db: AsyncSession = Depends(get_db),
):
    """Массовое создание задач в проекте в колонке 'К выполнению'."""
    await _get_project(project_id, db)
    if not data.tasks:
        raise HTTPException(status_code=400, detail="No tasks provided")

    created = await _bulk_create_tasks(project_id, data.tasks, data.attachments, db)
    broadcast({"type": "board_changed", "project_id": project_id})
    return TaskBulkOut(created=created, count=len(created))


@global_router.post("/kanban/tasks/bulk", response_model=TaskBulkOut, status_code=201)
async def create_tasks_bulk_global(
    data: TaskBulkRequest,
    db: AsyncSession = Depends(get_db),
):
    """Массовое создание задач в общем kanban. Каждая задача должна содержать project_id."""
    if not data.tasks:
        raise HTTPException(status_code=400, detail="No tasks provided")

    # Сгруппировать задачи по проектам и проверить существование проектов.
    tasks_by_project: dict[int, list[TaskBulkCreate]] = {}
    for task_data in data.tasks:
        if not task_data.project_id:
            raise HTTPException(status_code=400, detail="Each task must have project_id")
        tasks_by_project.setdefault(task_data.project_id, []).append(task_data)

    for pid in tasks_by_project:
        await _get_project(pid, db)

    created: list[Task] = []
    for project_id, project_tasks in tasks_by_project.items():
        project_created = await _bulk_create_tasks(project_id, project_tasks, data.attachments, db)
        created.extend(project_created)

    broadcast({"type": "board_changed", "global": True})
    return TaskBulkOut(created=created, count=len(created))


@global_router.post("/tasks/attachments/upload", response_model=TaskAttachmentOut, status_code=201)
async def upload_global_attachment_file(
    title: Optional[str] = Form(None),
    file: UploadFile = File(...),
):
    """Загрузить файл для последующего массового прикрепления к задачам (глобально)."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    uploads_dir = _task_uploads_dir()
    ext = Path(file.filename).suffix
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = uploads_dir / unique_name

    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)

    rel_path = f"tasks/{unique_name}"
    return TaskAttachmentOut(
        id=0,
        task_id=0,
        attachment_type="file",
        title=title or file.filename,
        file_path=rel_path,
        created_at=datetime.utcnow(),
    )


@global_router.post("/tasks/attachments/bulk", status_code=204)
async def add_attachments_to_selected_tasks_global(
    data: TaskBulkAddAttachmentsRequest,
    db: AsyncSession = Depends(get_db),
):
    """Добавить общие вложения к выбранным задачам (глобально, задачи из разных проектов)."""
    if not data.task_ids:
        raise HTTPException(status_code=400, detail="No task IDs provided")
    if not data.attachments:
        raise HTTPException(status_code=400, detail="No attachments provided")

    tasks_result = await db.execute(
        select(Task).where(Task.id.in_(data.task_ids))
    )
    tasks = tasks_result.scalars().all()
    found_ids = {t.id for t in tasks}
    missing = set(data.task_ids) - found_ids
    if missing:
        raise HTTPException(status_code=404, detail=f"Tasks not found: {sorted(missing)}")

    for task in tasks:
        for attachment_data in data.attachments:
            attachment = TaskAttachment(
                task_id=task.id,
                attachment_type=attachment_data.attachment_type,
                title=attachment_data.title,
                url=attachment_data.url,
                file_path=attachment_data.file_path,
            )
            db.add(attachment)

    await db.commit()
    for task in tasks:
        broadcast({"type": "task_changed", "project_id": task.project_id, "task_id": task.id, "status_id": task.status_id})
    return None


@global_router.get("/tasks", response_model=list[TaskOut])
async def list_tasks_global(
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    """Список всех задач (с возможностью фильтрации по проекту)."""
    query = select(Task).options(selectinload(Task.status), selectinload(Task.attachments))
    if project_id is not None:
        await _get_project(project_id, db)
        query = query.where(Task.project_id == project_id)
    result = await db.execute(query.order_by(Task.is_closed.asc(), Task.created_at.asc()))
    return result.scalars().all()


@global_router.get("/tasks/{task_id}", response_model=TaskOut)
async def get_task_global(task_id: int, db: AsyncSession = Depends(get_db)):
    """Получить задачу по глобальному ID (без указания проекта)."""
    result = await db.execute(
        select(Task)
        .options(selectinload(Task.status), selectinload(Task.attachments))
        .where(Task.id == task_id)
    )
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.patch("/tasks/reorder", status_code=204)
async def reorder_tasks(
    project_id: int,
    data: TaskReorderRequest,
    db: AsyncSession = Depends(get_db),
):
    await _get_project(project_id, db)
    await _get_status(project_id, data.status_id, db)

    old_statuses_result = await db.execute(
        select(Task.id, Task.status_id)
        .where(Task.id.in_(data.task_ids), Task.project_id == project_id)
    )
    old_statuses = {row.id: row.status_id for row in old_statuses_result.all()}

    for idx, task_id in enumerate(data.task_ids):
        await db.execute(
            update(Task)
            .where(Task.id == task_id, Task.project_id == project_id)
            .values(status_id=data.status_id, sort_order=idx)
        )
        old_status_id = old_statuses.get(task_id)
        if old_status_id != data.status_id:
            await _record_status_history(db, task_id, data.status_id)

    await db.commit()
    broadcast({"type": "board_changed", "project_id": project_id})
    return None


@router.get("/task-status-history")
async def get_task_status_history(
    project_id: int,
    task_ids: str = '',
    db: AsyncSession = Depends(get_db),
):
    """Получить историю переходов задач по колонкам."""
    await _get_project(project_id, db)
    ids = [int(x) for x in task_ids.split(',') if x.strip().isdigit()] if task_ids else []
    if not ids:
        return {"items": []}
    query = (
        select(TaskStatusHistory, TaskStatus.name)
        .join(TaskStatus, TaskStatusHistory.status_id == TaskStatus.id)
        .where(TaskStatusHistory.task_id.in_(ids))
        .order_by(TaskStatusHistory.entered_at.asc())
    )
    result = await db.execute(query)
    items = [
        TaskStatusHistoryOut(
            id=history.id,
            task_id=history.task_id,
            status_id=history.status_id,
            status_name=status_name,
            entered_at=history.entered_at,
        )
        for history, status_name in result.all()
    ]
    return {"items": items}


@router.get("/tasks/{task_id}", response_model=TaskOut)
async def get_task(
    project_id: int,
    task_id: int,
    db: AsyncSession = Depends(get_db),
):
    return await _get_task(project_id, task_id, db)


async def _apply_task_update(
    task: Task,
    data: TaskUpdate,
    project_id: int,
    db: AsyncSession,
) -> bool:
    """Применить поля TaskUpdate к одной задаче. Возвращает True, если статус изменился."""
    update_data = data.model_dump(exclude_unset=True)
    status_changed = False

    if "status_id" in update_data:
        await _get_status(project_id, update_data["status_id"], db)
        if task.status_id != update_data["status_id"]:
            task.status_id = update_data["status_id"]
            status_changed = True

    if "title" in update_data:
        task.title = update_data["title"] or ""
    if "priority" in update_data:
        task.priority = update_data["priority"]
    if "description" in update_data:
        task.description = update_data["description"]
    if "start_date" in update_data:
        task.start_date = update_data["start_date"]
    if "due_date" in update_data:
        task.due_date = update_data["due_date"]
    if "assignee_emails" in update_data or "assignee_email" in update_data:
        emails = _extract_assignee_emails(data)
        await _apply_task_assignees(task, emails, db)
    if "tags" in update_data:
        task.tags = update_data["tags"]
    if "list_name" in update_data:
        task.list_name = update_data["list_name"]
    if "result" in update_data:
        task.result = update_data["result"]

    if "is_closed" in update_data:
        was_closed = task.is_closed
        task.is_closed = update_data["is_closed"]
        if update_data["is_closed"] and not was_closed:
            task.due_date = None
            max_order = await db.execute(
                select(func.max(Task.sort_order))
                .where(Task.project_id == project_id, Task.status_id == task.status_id)
            )
            task.sort_order = (max_order.scalar_one_or_none() or 0) + 1
        elif not update_data["is_closed"] and was_closed:
            task.sort_order = 0
        elif status_changed:
            await db.execute(
                update(Task)
                .where(
                    Task.project_id == project_id,
                    Task.status_id == task.status_id,
                    Task.id != task.id,
                )
                .values(sort_order=Task.sort_order + 1)
            )
            task.sort_order = 0
    elif status_changed:
        await db.execute(
            update(Task)
            .where(
                Task.project_id == project_id,
                Task.status_id == task.status_id,
                Task.id != task.id,
            )
            .values(sort_order=Task.sort_order + 1)
        )
        task.sort_order = 0

    if status_changed:
        await _record_status_history(db, task.id, task.status_id)

    task.updated_at = datetime.utcnow()
    return status_changed


@router.patch("/tasks/{task_id}", response_model=TaskOut)
async def update_task(
    project_id: int,
    task_id: int,
    data: TaskUpdate,
    db: AsyncSession = Depends(get_db),
):
    task = await _get_task(project_id, task_id, db)
    old_result = task.result
    status_changed = await _apply_task_update(task, data, project_id, db)

    await db.commit()
    await db.refresh(task)
    broadcast({"type": "task_changed", "project_id": project_id, "task_id": task.id, "status_id": task.status_id})
    new_result = None
    update_data = data.model_dump(exclude_unset=True)
    if "result" in update_data:
        new_result = update_data["result"]
    if new_result and not old_result:
        broadcast({"type": "task_completed", "project_id": project_id, "task_id": task.id, "message": f"Задача №{task.id} выполнена"})
    return task


@router.patch("/tasks/bulk", status_code=204)
async def update_tasks_bulk(
    project_id: int,
    data: TaskBulkUpdateRequest,
    db: AsyncSession = Depends(get_db),
):
    """Массовое обновление задач проекта."""
    await _get_project(project_id, db)
    if not data.task_ids:
        raise HTTPException(status_code=400, detail="No task IDs provided")

    result = await db.execute(
        select(Task)
        .options(selectinload(Task.assignees))
        .where(Task.project_id == project_id, Task.id.in_(data.task_ids))
    )
    tasks = result.scalars().all()
    found_ids = {t.id for t in tasks}
    missing = set(data.task_ids) - found_ids
    if missing:
        raise HTTPException(status_code=404, detail=f"Tasks not found: {sorted(missing)}")

    for task in tasks:
        await _apply_task_update(task, data.update, project_id, db)

    await db.commit()
    for task in tasks:
        await db.refresh(task)
        broadcast({"type": "task_changed", "project_id": project_id, "task_id": task.id, "status_id": task.status_id})
    return None


@global_router.patch("/tasks/bulk", status_code=204)
async def update_tasks_bulk_global(
    data: TaskBulkUpdateRequest,
    db: AsyncSession = Depends(get_db),
):
    """Массовое обновление задач из общего списка (разные проекты).
    Смена статуса недоступна, т.к. колонки привязаны к проектам."""
    if not data.task_ids:
        raise HTTPException(status_code=400, detail="No task IDs provided")
    if data.update.status_id is not None:
        raise HTTPException(status_code=400, detail="status_id is not allowed in global bulk update")

    result = await db.execute(
        select(Task)
        .options(selectinload(Task.assignees))
        .where(Task.id.in_(data.task_ids))
    )
    tasks = result.scalars().all()
    found_ids = {t.id for t in tasks}
    missing = set(data.task_ids) - found_ids
    if missing:
        raise HTTPException(status_code=404, detail=f"Tasks not found: {sorted(missing)}")

    for task in tasks:
        await _apply_task_update(task, data.update, task.project_id, db)

    await db.commit()
    for task in tasks:
        await db.refresh(task)
        broadcast({"type": "task_changed", "project_id": task.project_id, "task_id": task.id, "status_id": task.status_id})
    return None


@router.delete("/tasks/{task_id}", status_code=204)
async def delete_task(
    project_id: int,
    task_id: int,
    db: AsyncSession = Depends(get_db),
):
    task = await _get_task(project_id, task_id, db)
    status_id = task.status_id
    await db.delete(task)
    await db.commit()
    broadcast({"type": "task_changed", "project_id": project_id, "task_id": task_id, "status_id": status_id, "deleted": True})
    return None


@router.get("/tasks/{task_id}/export", response_model=KanbanTaskExport)
async def export_single_task(
    project_id: int,
    task_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Экспорт одной задачи в формате KanbanTaskExport."""
    task = await _get_task(project_id, task_id, db)
    status = await _get_status(project_id, task.status_id, db)
    return {
        "id": task.id,
        "title": task.title,
        "description": task.description,
        "priority": task.priority,
        "is_closed": bool(task.is_closed),
        "start_date": task.start_date,
        "due_date": task.due_date,
        "assignee_email": task.assignee_email,
        "assignee_emails": task.assignee_emails,
        "tags": task.tags,
        "list_name": task.list_name,
        "result": task.result,
        "sort_order": task.sort_order,
        "status_name": status.name,
        "attachments": [
            {
                "attachment_type": a.attachment_type,
                "title": a.title,
                "url": a.url,
                "file_path": a.file_path,
            }
            for a in task.attachments
        ],
    }


@router.post("/tasks/import", response_model=TaskOut, status_code=201)
async def import_single_task(
    project_id: int,
    data: KanbanTaskExport,
    db: AsyncSession = Depends(get_db),
):
    """Импорт одной задачи. Если ID занят — создаётся задача с новым номером."""
    await _get_project(project_id, db)

    status_result = await db.execute(
        select(TaskStatus).where(
            TaskStatus.project_id == project_id,
            TaskStatus.name == data.status_name,
        )
    )
    status = status_result.scalar_one_or_none()
    if not status:
        max_order = await db.execute(
            select(func.max(TaskStatus.sort_order))
            .where(TaskStatus.project_id == project_id)
        )
        status = TaskStatus(
            project_id=project_id,
            name=data.status_name,
            color="#a78bfa",
            sort_order=(max_order.scalar_one_or_none() or 0) + 1,
        )
        db.add(status)
        await db.flush()
        await db.refresh(status)

    task_id_to_use = data.id
    if task_id_to_use:
        existing_result = await db.execute(
            select(Task).where(Task.project_id == project_id, Task.id == task_id_to_use)
        )
        if existing_result.scalar_one_or_none():
            task_id_to_use = None

    task_kwargs = {
        "project_id": project_id,
        "status_id": status.id,
        "title": data.title,
        "description": data.description,
        "priority": data.priority,
        "is_closed": data.is_closed,
        "start_date": data.start_date,
        "due_date": data.due_date,
        "assignee_email": None,
        "tags": data.tags,
        "list_name": data.list_name,
        "result": data.result,
        "sort_order": data.sort_order,
    }
    if task_id_to_use:
        task_kwargs["id"] = task_id_to_use

    task = Task(**task_kwargs)
    db.add(task)
    await db.flush()
    await db.refresh(task)

    emails = _extract_assignee_emails(data)
    await _apply_task_assignees(task, emails, db)

    for attachment_data in data.attachments or []:
        attachment = TaskAttachment(
            task_id=task.id,
            attachment_type=attachment_data.attachment_type,
            title=attachment_data.title,
            url=attachment_data.url,
            file_path=attachment_data.file_path,
        )
        db.add(attachment)

    await db.commit()
    await db.refresh(task, attribute_names=["status", "attachments"])
    return task


# ═══════════════════════════════════════════════════
# Вложения к задачам (файлы, ссылки, git)
# ═══════════════════════════════════════════════════


def _task_uploads_dir() -> Path:
    settings = get_settings()
    base = Path(settings.local_storage_path).resolve()
    uploads = base / "tasks"
    uploads.mkdir(parents=True, exist_ok=True)
    return uploads


@router.post("/attachments/upload", response_model=TaskAttachmentOut, status_code=201)
async def upload_bulk_attachment_file(
    project_id: int,
    title: Optional[str] = Form(None),
    file: UploadFile = File(...),
):
    """Загрузить файл для последующего массового прикрепления к задачам."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    uploads_dir = _task_uploads_dir()
    ext = Path(file.filename).suffix
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = uploads_dir / unique_name

    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)

    rel_path = f"tasks/{unique_name}"
    return TaskAttachmentOut(
        id=0,
        task_id=0,
        attachment_type="file",
        title=title or file.filename,
        file_path=rel_path,
        created_at=datetime.utcnow(),
    )


@router.post("/tasks/{task_id}/attachments", response_model=TaskAttachmentOut, status_code=201)
async def add_task_attachment(
    project_id: int,
    task_id: int,
    data: TaskAttachmentCreate,
    db: AsyncSession = Depends(get_db),
):
    task = await _get_task(project_id, task_id, db)
    attachment = TaskAttachment(
        task_id=task.id,
        attachment_type=data.attachment_type,
        title=data.title,
        url=data.url,
        file_path=data.file_path,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)
    return attachment


@router.post("/tasks/{task_id}/attachments/upload", response_model=TaskAttachmentOut, status_code=201)
async def upload_task_attachment_file(
    project_id: int,
    task_id: int,
    title: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    task = await _get_task(project_id, task_id, db)
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    uploads_dir = _task_uploads_dir()
    ext = Path(file.filename).suffix
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = uploads_dir / unique_name

    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)

    rel_path = f"tasks/{unique_name}"
    attachment = TaskAttachment(
        task_id=task.id,
        attachment_type="file",
        title=title or file.filename,
        file_path=rel_path,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)
    return attachment


@router.delete("/tasks/{task_id}/attachments/{attachment_id}", status_code=204)
async def delete_task_attachment(
    project_id: int,
    task_id: int,
    attachment_id: int,
    db: AsyncSession = Depends(get_db),
):
    task = await _get_task(project_id, task_id, db)
    result = await db.execute(
        select(TaskAttachment).where(
            TaskAttachment.id == attachment_id,
            TaskAttachment.task_id == task.id,
        )
    )
    attachment = result.scalar_one_or_none()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")

    if attachment.attachment_type == "file" and attachment.file_path:
        try:
            full_path = Path(get_settings().local_storage_path).resolve() / attachment.file_path
            if full_path.exists():
                full_path.unlink()
        except OSError:
            pass

    await db.delete(attachment)
    await db.commit()
    return None


@router.get("/tasks/{task_id}/attachments/{attachment_id}/collabora")
async def collabora_task_attachment(
    project_id: int,
    task_id: int,
    attachment_id: int,
    db: AsyncSession = Depends(get_db),
    storage: StorageBackend = Depends(get_storage),
    settings: Settings = Depends(get_settings),
):
    """Вернуть URL iframe для редактирования вложения задачи в Collabora Online."""
    if not settings.collabora_enabled:
        raise HTTPException(status_code=503, detail="Collabora Online is not enabled")

    task = await _get_task(project_id, task_id, db)
    result = await db.execute(
        select(TaskAttachment).where(
            TaskAttachment.id == attachment_id,
            TaskAttachment.task_id == task.id,
        )
    )
    attachment = result.scalar_one_or_none()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")

    if attachment.attachment_type != "file" or not attachment.file_path:
        raise HTTPException(status_code=400, detail="Not a file attachment")

    ext = Path(attachment.title or attachment.file_path).suffix.lower()
    if ext not in OFFICE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported file type for Collabora")

    local_path = await storage.get_local_path(attachment.file_path)
    if not local_path or not os.path.exists(local_path):
        raise HTTPException(status_code=400, detail="File is not available locally")

    file_id = encode_file_id(project_file_path=attachment.file_path)
    iframe_url = await build_collabora_iframe_url(file_id, ext, settings)
    return {
        "url": iframe_url,
        "wopi_src": f"{settings.collabora_wopi_internal_url.rstrip('/')}/wopi/files/{file_id}",
    }


@router.post("/tasks/{task_id}/attachments/{attachment_id}/open")
async def open_task_attachment(
    request: Request,
    project_id: int,
    task_id: int,
    attachment_id: int,
    db: AsyncSession = Depends(get_db),
    storage: StorageBackend = Depends(get_storage),
):
    """Открыть файл-вложение задачи в системном приложении. Только localhost."""
    task = await _get_task(project_id, task_id, db)
    result = await db.execute(
        select(TaskAttachment).where(
            TaskAttachment.id == attachment_id,
            TaskAttachment.task_id == task.id,
        )
    )
    attachment = result.scalar_one_or_none()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")

    if attachment.attachment_type != "file" or not attachment.file_path:
        raise HTTPException(status_code=400, detail="Not a file attachment")

    client_host = request.client.host if request.client else None
    if client_host not in ("127.0.0.1", "::1", "localhost"):
        raise HTTPException(status_code=403, detail="Remote open not allowed")

    local_path = await storage.get_local_path(attachment.file_path)
    if not local_path or not os.path.exists(local_path):
        raise HTTPException(status_code=404, detail="File not found")

    def _get_gui_env():
        env = os.environ.copy()
        uid = str(os.getuid())
        runtime_dir = f"/run/user/{uid}"
        env["XDG_RUNTIME_DIR"] = runtime_dir
        env["DBUS_SESSION_BUS_ADDRESS"] = f"unix:path={runtime_dir}/bus"
        display = None
        try:
            result = subprocess.run(
                ["pgrep", "-a", "-u", uid, "-f", "Xwayland|Xorg"],
                capture_output=True, text=True, timeout=5
            )
            for line in result.stdout.strip().split("\n"):
                if not line:
                    continue
                parts = line.split()
                for i, part in enumerate(parts):
                    if part in ("Xwayland", "Xorg") and i + 1 < len(parts):
                        candidate = parts[i + 1]
                        if candidate.startswith(":"):
                            display = candidate
                            break
                if display:
                    break
        except Exception:
            pass
        if not display:
            try:
                result = subprocess.run(
                    ["ps", "e", "-u", uid],
                    capture_output=True, text=True, timeout=5
                )
                for line in result.stdout.split("\n"):
                    if "DISPLAY=:" in line:
                        for part in line.split():
                            if part.startswith("DISPLAY=:"):
                                display = part.split("=", 1)[1]
                                break
                    if display:
                        break
            except Exception:
                pass
        if display:
            env["DISPLAY"] = display

        xauthority = None
        try:
            result = subprocess.run(
                ["ps", "e", "-u", uid],
                capture_output=True, text=True, timeout=5
            )
            for line in result.stdout.split("\n"):
                if "XAUTHORITY=" in line:
                    for part in line.split():
                        if part.startswith("XAUTHORITY="):
                            xauthority = part.split("=", 1)[1]
                            break
                if xauthority:
                    break
        except Exception:
            pass
        if xauthority:
            env["XAUTHORITY"] = xauthority

        return env

    system = platform.system()
    try:
        if system == "Windows":
            os.startfile(local_path)
        elif system == "Darwin":
            subprocess.Popen(["open", local_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        else:
            gui_env = _get_gui_env()
            subprocess.Popen(
                ["xdg-open", local_path],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                env=gui_env,
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open file: {str(e)}")

    return {"status": "opened", "path": local_path}


@router.patch("/tasks/{task_id}/attachments/{attachment_id}", response_model=TaskAttachmentOut)
async def update_task_attachment(
    project_id: int,
    task_id: int,
    attachment_id: int,
    title: Optional[str] = Form(None),
    url: Optional[str] = Form(None),
    file: UploadFile = File(None),
    db: AsyncSession = Depends(get_db),
):
    task = await _get_task(project_id, task_id, db)
    result = await db.execute(
        select(TaskAttachment).where(
            TaskAttachment.id == attachment_id,
            TaskAttachment.task_id == task.id,
        )
    )
    attachment = result.scalar_one_or_none()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")

    if title is not None:
        attachment.title = title.strip() or None

    if attachment.attachment_type in ("link", "git"):
        if url is not None:
            attachment.url = url.strip() or None
    elif file and file.filename:
        uploads_dir = _task_uploads_dir()
        ext = Path(file.filename).suffix
        unique_name = f"{uuid.uuid4().hex}{ext}"
        file_path = uploads_dir / unique_name

        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)

        if attachment.file_path:
            try:
                old_full_path = Path(get_settings().local_storage_path).resolve() / attachment.file_path
                if old_full_path.exists():
                    old_full_path.unlink()
            except OSError:
                pass

        attachment.file_path = f"tasks/{unique_name}"
        if not attachment.title:
            attachment.title = file.filename

    await db.commit()
    await db.refresh(attachment)
    return attachment


@router.post("/tasks/attachments/bulk", status_code=204)
async def add_attachments_to_selected_tasks(
    project_id: int,
    data: TaskBulkAddAttachmentsRequest,
    db: AsyncSession = Depends(get_db),
):
    """Добавить общие вложения к выбранным задачам проекта."""
    await _get_project(project_id, db)
    if not data.task_ids:
        raise HTTPException(status_code=400, detail="No task IDs provided")
    if not data.attachments:
        raise HTTPException(status_code=400, detail="No attachments provided")

    tasks_result = await db.execute(
        select(Task).where(Task.project_id == project_id, Task.id.in_(data.task_ids))
    )
    tasks = tasks_result.scalars().all()
    found_ids = {t.id for t in tasks}
    missing = set(data.task_ids) - found_ids
    if missing:
        raise HTTPException(status_code=404, detail=f"Tasks not found: {sorted(missing)}")

    for task in tasks:
        for attachment_data in data.attachments:
            attachment = TaskAttachment(
                task_id=task.id,
                attachment_type=attachment_data.attachment_type,
                title=attachment_data.title,
                url=attachment_data.url,
                file_path=attachment_data.file_path,
            )
            db.add(attachment)

    await db.commit()
    for task in tasks:
        broadcast({"type": "task_changed", "project_id": project_id, "task_id": task.id, "status_id": task.status_id})
    return None


# ═══════════════════════════════════════════════════
# Экспорт / импорт kanban проекта
# ═══════════════════════════════════════════════════

async def _build_project_export(project_id: int, db: AsyncSession) -> dict:
    """Собрать данные для экспорта одного проекта."""
    project = await _get_project(project_id, db)

    statuses_result = await db.execute(
        select(TaskStatus)
        .where(TaskStatus.project_id == project_id)
        .order_by(TaskStatus.sort_order.asc(), TaskStatus.created_at.asc())
    )
    statuses = statuses_result.scalars().all()

    tasks_result = await db.execute(
        select(Task)
        .options(selectinload(Task.attachments))
        .where(Task.project_id == project_id)
        .order_by(Task.is_closed.asc(), Task.sort_order.asc(), Task.created_at.asc())
    )
    tasks = tasks_result.scalars().all()

    status_map = {s.id: s.name for s in statuses}

    return {
        "name": project.name,
        "description": project.description,
        "statuses": [
            {
                "name": s.name,
                "color": s.color,
                "sort_order": s.sort_order,
            }
            for s in statuses
        ],
        "tasks": [
            {
                "id": t.id,
                "title": t.title,
                "description": t.description,
                "priority": t.priority,
                "is_closed": t.is_closed,
                "start_date": t.start_date,
                "due_date": t.due_date,
                "assignee_email": t.assignee_email,
                "assignee_emails": t.assignee_emails,
                "tags": t.tags,
                "list_name": t.list_name,
                "sort_order": t.sort_order,
                "status_name": status_map.get(t.status_id, ""),
                "attachments": [
                    {
                        "attachment_type": a.attachment_type,
                        "title": a.title,
                        "url": a.url,
                        "file_path": a.file_path,
                    }
                    for a in t.attachments
                ],
            }
            for t in tasks
        ],
    }


async def _import_project_data(
    project_id: int,
    project_data: KanbanProjectExport,
    db: AsyncSession,
) -> tuple[int, int]:
    """Импортировать статусы и задачи в указанный проект."""
    imported_statuses = 0
    imported_tasks = 0

    status_name_to_id = {}
    for status_data in project_data.statuses:
        status_result = await db.execute(
            select(TaskStatus).where(
                TaskStatus.project_id == project_id,
                TaskStatus.name == status_data.name,
            )
        )
        status = status_result.scalar_one_or_none()
        if not status:
            status = TaskStatus(
                project_id=project_id,
                name=status_data.name,
                color=status_data.color,
                sort_order=status_data.sort_order,
            )
            db.add(status)
            await db.flush()
            await db.refresh(status)
            imported_statuses += 1
        else:
            status.color = status_data.color
            status.sort_order = status_data.sort_order
        status_name_to_id[status.name] = status.id

    for task_data in project_data.tasks:
        status_id = status_name_to_id.get(task_data.status_name)
        if not status_id:
            status_result = await db.execute(
                select(TaskStatus).where(
                    TaskStatus.project_id == project_id,
                    TaskStatus.name == task_data.status_name,
                )
            )
            status = status_result.scalar_one_or_none()
            if not status:
                continue
            status_id = status.id

        task_result = await db.execute(
            select(Task).where(
                Task.project_id == project_id,
                Task.title == task_data.title,
                Task.status_id == status_id,
            )
        )
        task = task_result.scalar_one_or_none()
        if not task:
            task = Task(
                project_id=project_id,
                status_id=status_id,
                title=task_data.title,
                description=task_data.description,
                priority=task_data.priority,
                is_closed=task_data.is_closed,
                start_date=task_data.start_date,
                due_date=task_data.due_date,
                assignee_email=None,
                tags=task_data.tags,
                list_name=task_data.list_name,
                sort_order=task_data.sort_order,
            )
            db.add(task)
            imported_tasks += 1
        else:
            task.description = task_data.description
            task.priority = task_data.priority
            task.is_closed = task_data.is_closed
            task.start_date = task_data.start_date
            task.due_date = task_data.due_date
            task.assignee_email = None
            task.tags = task_data.tags
            task.list_name = task_data.list_name
            task.sort_order = task_data.sort_order

        await db.flush()
        await db.refresh(task)

        emails = _extract_assignee_emails(task_data)
        await _apply_task_assignees(task, emails, db)

        attachments_to_import = list(task_data.attachments or [])
        if attachments_to_import:
            old_attachments_result = await db.execute(
                select(TaskAttachment).where(TaskAttachment.task_id == task.id)
            )
            for old in old_attachments_result.scalars().all():
                if old.attachment_type == "file" and old.file_path:
                    try:
                        full_path = Path(get_settings().local_storage_path).resolve() / old.file_path
                        if full_path.exists():
                            full_path.unlink()
                    except OSError:
                        pass
                await db.delete(old)

            for attachment_data in attachments_to_import:
                attachment = TaskAttachment(
                    task_id=task.id,
                    attachment_type=attachment_data.attachment_type,
                    title=attachment_data.title,
                    url=attachment_data.url,
                    file_path=attachment_data.file_path,
                )
                db.add(attachment)

    return imported_statuses, imported_tasks


@router.get("/kanban/export", response_model=KanbanExportOut)
async def export_project_kanban(project_id: int, db: AsyncSession = Depends(get_db)):
    """Экспорт колонок и задач kanban конкретного проекта."""
    project_export = await _build_project_export(project_id, db)
    return {
        "version": 1,
        "exported_at": datetime.utcnow(),
        "projects": [project_export],
    }


@router.post("/kanban/import")
async def import_project_kanban(
    project_id: int,
    data: KanbanImportIn,
    db: AsyncSession = Depends(get_db),
):
    """Импорт колонок и задач kanban в конкретный проект."""
    await _get_project(project_id, db)

    imported_statuses = 0
    imported_tasks = 0

    for project_data in data.projects:
        statuses, tasks = await _import_project_data(project_id, project_data, db)
        imported_statuses += statuses
        imported_tasks += tasks

    await db.commit()
    return {
        "success": True,
        "imported_statuses": imported_statuses,
        "imported_tasks": imported_tasks,
    }


# ═══════════════════════════════════════════════════
# Общий kanban
# ═══════════════════════════════════════════════════

@global_router.get("/kanban", response_model=KanbanGlobalOut)
async def global_kanban(
    project_id: Optional[int] = None,
    priority: Optional[str] = None,
    assignee_email: Optional[str] = None,
    tags: Optional[str] = None,
    created_after: Optional[datetime] = None,
    created_before: Optional[datetime] = None,
    due_after: Optional[datetime] = None,
    due_before: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
):
    """Все задачи всех проектов с фильтрами для общего kanban."""
    query = select(Task).options(selectinload(Task.status), selectinload(Task.attachments))

    if project_id is not None:
        query = query.where(Task.project_id == project_id)
    if priority is not None:
        query = query.where(Task.priority == priority)
    if assignee_email is not None:
        search = f"%{assignee_email}%"
        query = query.outerjoin(
            TaskAssignee, Task.id == TaskAssignee.task_id
        ).where(
            or_(
                Task.assignee_email.ilike(search),
                TaskAssignee.assignee_email.ilike(search),
            )
        ).distinct()
    if tags is not None:
        query = query.where(Task.tags.ilike(f"%{tags}%"))
    if created_after is not None:
        query = query.where(Task.created_at >= created_after)
    if created_before is not None:
        query = query.where(Task.created_at <= created_before)
    if due_after is not None:
        query = query.where(Task.due_date >= due_after)
    if due_before is not None:
        query = query.where(Task.due_date <= due_before)

    result = await db.execute(query.order_by(Task.is_closed.asc(), Task.sort_order.asc(), Task.created_at.asc()))
    tasks = result.scalars().all()

    # Колонки — уникальные названия статусов с приоритетным цветом (самый частый)
    color_query = select(TaskStatus.name, TaskStatus.color, func.count(TaskStatus.id).label("cnt"))
    if project_id is not None:
        color_query = color_query.where(TaskStatus.project_id == project_id)
    color_query = color_query.group_by(TaskStatus.name, TaskStatus.color)
    color_result = await db.execute(color_query)

    column_colors = {}
    for name, color, cnt in color_result.fetchall():
        if name not in column_colors or cnt > column_colors[name]["cnt"]:
            column_colors[name] = {"color": color, "cnt": cnt}

    # Порядок колонок определяется минимальным sort_order среди статусов с таким именем
    order_query = select(TaskStatus.name, func.min(TaskStatus.sort_order).label("min_order"))
    if project_id is not None:
        order_query = order_query.where(TaskStatus.project_id == project_id)
    order_query = order_query.group_by(TaskStatus.name)
    order_result = await db.execute(order_query)
    column_order = {name: min_order for name, min_order in order_result.fetchall()}

    columns = [KanbanColumnOut(name=name, color=data["color"]) for name, data in column_colors.items()]
    columns.sort(key=lambda c: (column_order.get(c.name, 0), c.name))

    return KanbanGlobalOut(columns=columns, tasks=tasks)


@global_router.get("/kanban/filters", response_model=KanbanFiltersOut)
async def global_kanban_filters(db: AsyncSession = Depends(get_db)):
    """Доступные значения фильтров для общего kanban."""
    projects_result = await db.execute(select(Project).order_by(Project.name.asc()))
    projects = projects_result.scalars().all()

    priorities_result = await db.execute(select(distinct(Task.priority)).where(Task.priority.isnot(None)))
    priorities = [p[0] for p in priorities_result.fetchall() if p[0]]

    assignees_result = await db.execute(select(Assignee).order_by(Assignee.name.asc()))
    assignees = assignees_result.scalars().all()

    tags_result = await db.execute(select(Task.tags).where(Task.tags.isnot(None)))
    tag_set = set()
    for (tags_str,) in tags_result.fetchall():
        for tag in tags_str.split(","):
            tag = tag.strip()
            if tag:
                tag_set.add(tag)

    list_names_result = await db.execute(
        select(distinct(Task.list_name)).where(Task.list_name.isnot(None))
    )
    list_names = [ln[0] for ln in list_names_result.fetchall() if ln[0]]

    return KanbanFiltersOut(
        projects=projects,
        priorities=sorted(priorities),
        assignees=assignees,
        tags=sorted(tag_set),
        list_names=sorted(list_names),
    )


@global_router.get("/kanban/columns", response_model=list[KanbanColumnOut])
async def list_global_kanban_columns(db: AsyncSession = Depends(get_db)):
    """Уникальные колонки общего kanban с самым частым цветом."""
    color_query = select(TaskStatus.name, TaskStatus.color, func.count(TaskStatus.id).label("cnt")).group_by(
        TaskStatus.name, TaskStatus.color
    )
    color_result = await db.execute(color_query)

    column_colors = {}
    for name, color, cnt in color_result.fetchall():
        if name not in column_colors or cnt > column_colors[name]["cnt"]:
            column_colors[name] = {"color": color, "cnt": cnt}

    order_query = select(TaskStatus.name, func.min(TaskStatus.sort_order).label("min_order")).group_by(
        TaskStatus.name
    )
    order_result = await db.execute(order_query)
    column_order = {name: min_order for name, min_order in order_result.fetchall()}

    columns = [KanbanColumnOut(name=name, color=data["color"]) for name, data in column_colors.items()]
    columns.sort(key=lambda c: (column_order.get(c.name, 0), c.name))
    return columns


@global_router.post("/kanban/columns", response_model=list[TaskStatusOut], status_code=201)
async def create_global_kanban_column(
    data: GlobalKanbanColumnCreate,
    db: AsyncSession = Depends(get_db),
):
    """Создать колонку с указанным именем во всех проектах, где её ещё нет."""
    projects_result = await db.execute(select(Project))
    projects = projects_result.scalars().all()
    if not projects:
        raise HTTPException(status_code=400, detail="No projects found")

    created = []
    for project in projects:
        existing = await db.execute(
            select(TaskStatus).where(TaskStatus.project_id == project.id, TaskStatus.name == data.name)
        )
        if existing.scalar_one_or_none():
            continue

        max_order = await db.execute(
            select(TaskStatus.sort_order)
            .where(TaskStatus.project_id == project.id)
            .order_by(TaskStatus.sort_order.desc())
            .limit(1)
        )
        max_val = max_order.scalar_one_or_none() or 0
        status = TaskStatus(
            project_id=project.id,
            name=data.name,
            color=data.color or "#a78bfa",
            sort_order=max_val + 1,
        )
        db.add(status)
        await db.flush()
        await db.refresh(status)
        created.append(status)

    await db.commit()
    broadcast({"type": "board_changed", "global": True})
    return created


@global_router.patch("/kanban/columns/{name}", response_model=list[TaskStatusOut])
async def update_global_kanban_column(
    name: str,
    data: GlobalKanbanColumnUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Переименовать или перекрасить колонку во всех проектах."""
    result = await db.execute(select(TaskStatus).where(TaskStatus.name == name))
    statuses = result.scalars().all()
    if not statuses:
        raise HTTPException(status_code=404, detail="Column not found")

    new_name = data.new_name.strip() if data.new_name else None
    if new_name:
        for status in statuses:
            conflict = await db.execute(
                select(TaskStatus).where(
                    TaskStatus.project_id == status.project_id,
                    TaskStatus.name == new_name,
                    TaskStatus.id != status.id,
                )
            )
            if conflict.scalar_one_or_none():
                raise HTTPException(
                    status_code=400,
                    detail=f"Column '{new_name}' already exists in project",
                )
            status.name = new_name

    if data.color:
        for status in statuses:
            status.color = data.color

    await db.commit()
    for status in statuses:
        await db.refresh(status)
    broadcast({"type": "board_changed", "global": True})
    return statuses


@global_router.delete("/kanban/columns/{name}", status_code=204)
async def delete_global_kanban_column(
    name: str,
    db: AsyncSession = Depends(get_db),
):
    """Удалить колонку во всех проектах, если в них нет задач."""
    result = await db.execute(select(TaskStatus).where(TaskStatus.name == name))
    statuses = result.scalars().all()
    if not statuses:
        raise HTTPException(status_code=404, detail="Column not found")

    for status in statuses:
        tasks_result = await db.execute(select(Task.id).where(Task.status_id == status.id).limit(1))
        if tasks_result.scalar_one_or_none():
            raise HTTPException(
                status_code=400,
                detail=f"Column '{name}' has tasks in some projects",
            )

    for status in statuses:
        await db.delete(status)
    await db.commit()
    broadcast({"type": "board_changed", "global": True})
    return None


@global_router.patch("/kanban/tasks/{task_id}/status", response_model=TaskOut)
async def update_task_status_by_column_name(
    task_id: int,
    data: KanbanTaskStatusUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Обновить статус задачи, найдя статус проекта по названию колонки."""
    result = await db.execute(
        select(Task)
        .options(selectinload(Task.status), selectinload(Task.attachments))
        .where(Task.id == task_id)
    )
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    status_result = await db.execute(
        select(TaskStatus)
        .where(TaskStatus.project_id == task.project_id, TaskStatus.name == data.column_name)
        .order_by(TaskStatus.sort_order.asc())
        .limit(1)
    )
    status = status_result.scalar_one_or_none()
    if not status:
        raise HTTPException(
            status_code=400,
            detail=f"Status '{data.column_name}' not found in project"
        )

    if data.insert_top and task.status_id != status.id:
        await db.execute(
            update(Task)
            .where(
                Task.project_id == task.project_id,
                Task.status_id == status.id,
                Task.id != task.id,
            )
            .values(sort_order=Task.sort_order + 1)
        )
        task.sort_order = 0

    status_changed = task.status_id != status.id
    task.status_id = status.id
    task.updated_at = datetime.utcnow()
    if status_changed:
        await _record_status_history(db, task.id, status.id)
    await db.commit()
    await db.refresh(task)
    broadcast({"type": "board_changed", "project_id": task.project_id, "global": True})
    return task


@global_router.get("/task-status-history")
async def get_global_task_status_history(
    task_ids: str = '',
    db: AsyncSession = Depends(get_db),
):
    """Получить историю переходов задач по колонкам (глобально)."""
    ids = [int(x) for x in task_ids.split(',') if x.strip().isdigit()] if task_ids else []
    if not ids:
        return {"items": []}
    query = (
        select(TaskStatusHistory, TaskStatus.name)
        .join(TaskStatus, TaskStatusHistory.status_id == TaskStatus.id)
        .where(TaskStatusHistory.task_id.in_(ids))
        .order_by(TaskStatusHistory.entered_at.asc())
    )
    result = await db.execute(query)
    items = [
        TaskStatusHistoryOut(
            id=history.id,
            task_id=history.task_id,
            status_id=history.status_id,
            status_name=status_name,
            entered_at=history.entered_at,
        )
        for history, status_name in result.all()
    ]
    return {"items": items}


@global_router.post("/tasks/telegram-list-config", status_code=204)
async def save_telegram_task_list_config(data: TaskListTelegramConfig):
    """Сохранить конфигурацию списка задач для отправки в Telegram."""
    config_path = Path("data/telegram_task_list.json")
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(data.model_dump_json(indent=2), encoding="utf-8")
    return None


@global_router.get("/kanban/export", response_model=KanbanExportOut)
async def export_global_kanban(db: AsyncSession = Depends(get_db)):
    """Экспорт всех колонок и задач kanban по проектам."""
    projects_result = await db.execute(select(Project).order_by(Project.name.asc()))
    projects = projects_result.scalars().all()

    export_projects = []
    for project in projects:
        export_projects.append(await _build_project_export(project.id, db))

    return {
        "version": 1,
        "exported_at": datetime.utcnow(),
        "projects": export_projects,
    }


@global_router.post("/kanban/import")
async def import_global_kanban(
    data: KanbanImportIn,
    db: AsyncSession = Depends(get_db),
):
    """Импорт колонок и задач kanban из JSON-дампа."""
    imported_projects = 0
    imported_statuses = 0
    imported_tasks = 0

    for project_data in data.projects:
        project_result = await db.execute(
            select(Project).where(Project.name == project_data.name)
        )
        project = project_result.scalar_one_or_none()
        if not project:
            project = Project(
                name=project_data.name,
                description=project_data.description,
            )
            db.add(project)
            await db.flush()
            await db.refresh(project)
        imported_projects += 1

        statuses, tasks = await _import_project_data(project.id, project_data, db)
        imported_statuses += statuses
        imported_tasks += tasks

    await db.commit()
    broadcast({"type": "board_changed", "global": True})
    return {
        "success": True,
        "imported_projects": imported_projects,
        "imported_statuses": imported_statuses,
        "imported_tasks": imported_tasks,
    }


# ═══════════════════════════════════════════════════
# Экспорт диаграммы Ганта в Excel
# ═══════════════════════════════════════════════════

@global_router.get("/gantt/export/xlsx")
async def export_gantt_xlsx(
    project_id: Optional[int] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assignee_email: Optional[str] = None,
    list_name: Optional[str] = None,
    closed: Optional[int] = None,
    tags: Optional[str] = None,
    hide_no_deadline: bool = False,
    status_overlay: bool = False,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "asc",
    db: AsyncSession = Depends(get_db),
):
    """Экспорт диаграммы Ганта в .xlsx с учётом фильтров."""
    try:
        query = select(Task).options(
            selectinload(Task.status),
            selectinload(Task.project),
            selectinload(Task.assignees),
        )

        if project_id is not None:
            await _get_project(project_id, db)
            query = query.where(Task.project_id == project_id)
        if priority is not None:
            query = query.where(Task.priority == priority)
        if assignee_email is not None:
            search = f"%{assignee_email}%"
            query = query.outerjoin(
                TaskAssignee, Task.id == TaskAssignee.task_id
            ).where(
                or_(
                    Task.assignee_email.ilike(search),
                    TaskAssignee.assignee_email.ilike(search),
                )
            ).distinct()
        if list_name is not None:
            query = query.where(Task.list_name.ilike(f"%{list_name}%"))
        if closed is not None:
            query = query.where(Task.is_closed == bool(closed))
        if status is not None:
            query = query.join(TaskStatus, Task.status_id == TaskStatus.id)
            query = query.where(TaskStatus.name.ilike(f"%{status}%"))
        if search is not None:
            search_lower = f"%{search.lower()}%"
            query = query.where(
                func.lower(Task.title).ilike(search_lower)
                | func.lower(Task.description).ilike(search_lower)
            )
        if tags is not None:
            tag_list = [t.strip().lower() for t in tags.split(",") if t.strip()]
            for tag in tag_list:
                query = query.where(func.lower(Task.tags).ilike(f"%{tag}%"))
        if hide_no_deadline:
            query = query.where(Task.due_date.isnot(None))

        sort_column_map = {
            "created_at": Task.created_at,
            "due_date": Task.due_date,
            "priority": Task.priority,
            "title": Task.title,
            "id": Task.id,
            "updated_at": Task.updated_at,
            "assignee_name": Task.assignee_email,
        }
        if sort_by == "status_name":
            query = query.join(TaskStatus, Task.status_id == TaskStatus.id)
            order_column = TaskStatus.name
        else:
            order_column = sort_column_map.get(sort_by, Task.created_at)

        is_desc = str(sort_order or "asc").lower() == "desc"
        query = query.order_by(order_column.desc() if is_desc else order_column.asc())

        result = await db.execute(query)
        tasks = result.scalars().all()

        if not tasks:
            raise HTTPException(status_code=404, detail="Нет задач для экспорта")

        # Карта имён ответственных
        assignees_result = await db.execute(select(Assignee))
        assignees = {a.email: a.name for a in assignees_result.scalars().all()}

        # Статусы Kanban для оверлея: 3-я и 4-я колонки
        status_overlay_map = {}
        if status_overlay:
            project_ids = list({t.project_id for t in tasks})
            for pid in project_ids:
                statuses_result = await db.execute(
                    select(TaskStatus)
                    .where(TaskStatus.project_id == pid)
                    .order_by(TaskStatus.sort_order.asc())
                )
                project_statuses = statuses_result.scalars().all()
                if len(project_statuses) > 2:
                    status_overlay_map[pid] = {
                        "testing": project_statuses[2].id,
                        "deploy": project_statuses[3].id if len(project_statuses) > 3 else None,
                    }

            task_ids = [t.id for t in tasks]
            history_by_task = {}
            if task_ids:
                history_result = await db.execute(
                    select(TaskStatusHistory)
                    .where(TaskStatusHistory.task_id.in_(task_ids))
                    .order_by(TaskStatusHistory.entered_at.asc())
                )
                for record in history_result.scalars().all():
                    history_by_task.setdefault(record.task_id, []).append(record)

            def get_first_status_date(task_id: int, status_id: Optional[int]):
                if not status_id:
                    return None
                for record in history_by_task.get(task_id, []):
                    if record.status_id == status_id and record.entered_at:
                        return record.entered_at.date()
                return None
        else:
            def get_first_status_date(task_id: int, status_id: Optional[int]):
                return None

        # Даты
        def date_part(dt):
            if not dt:
                return None
            return dt.date() if hasattr(dt, "date") else dt

        all_dates = []
        for t in tasks:
            s = date_part(t.start_date) or date_part(t.created_at)
            e = date_part(t.due_date)
            if s:
                all_dates.append(s)
            if e:
                all_dates.append(e)

        if not all_dates:
            raise HTTPException(status_code=400, detail="У отфильтрованных задач отсутствуют даты")

        min_date = min(all_dates) - timedelta(days=2)
        max_date = max(all_dates) + timedelta(days=2)
        total_days = (max_date - min_date).days + 1

        wb = Workbook()
        wb.remove(wb.active)

        # ── Лист Гант ──
        ws = wb.create_sheet("Гант")

        header_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        today_fill = PatternFill(start_color="DEE2E6", end_color="DEE2E6", fill_type="solid")
        fills = {
            "low": PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid"),
            "medium": PatternFill(start_color="FFECB5", end_color="FFECB5", fill_type="solid"),
            "high": PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid"),
            "closed": PatternFill(start_color="D6D8DB", end_color="D6D8DB", fill_type="solid"),
        }

        thin_border_side = Side(style="thin", color="DEE2E6")
        thin_border = Border(
            left=thin_border_side, right=thin_border_side,
            top=thin_border_side, bottom=thin_border_side
        )

        # Заголовки
        headers = ["#", "Название", "Исполнитель", "Начало", "Конец", "Длит."]
        day_headers = []
        day_dates = []
        for i in range(total_days):
            d = min_date + timedelta(days=i)
            day_dates.append(d)
            month = d.strftime("%b").lower().rstrip(".")
            day_headers.append(f"{month}\n{d.day}")
        ws.append(headers + day_headers)

        today = datetime.utcnow().date()

        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill = header_fill
            cell.font = Font(bold=True, color="212529")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx > 6:
                cell.number_format = "DD.MM"
                ws.column_dimensions[cell.column_letter].width = 4.5

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 10

        for t in tasks:
            start = date_part(t.start_date) or date_part(t.created_at)
            end = date_part(t.due_date)
            if not start and not end:
                continue
            if not start:
                start = end

            title = (t.title or "").strip() or (t.description or "").strip() or "—"
            assignee_str = _format_task_assignees(t, assignees)
            start_str = start.strftime("%d.%m.%Y") if start else ""
            end_str = end.strftime("%d.%m.%Y") if end else ""
            duration = (end - start).days + 1 if end else ""

            testing_date = None
            deploy_date = None
            if status_overlay:
                mapping = status_overlay_map.get(t.project_id, {})
                testing_date = get_first_status_date(t.id, mapping.get("testing"))
                deploy_date = get_first_status_date(t.id, mapping.get("deploy"))

            row_cells = [t.id, title, assignee_str, start_str, end_str, duration]
            ws.append(row_cells)
            row_num = ws.max_row

            fill = fills["closed"] if t.is_closed else fills.get(t.priority, fills["medium"])

            for i, d in enumerate(day_dates):
                col_num = 7 + i
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                is_active = False
                if end:
                    visible_start = max(start, min_date)
                    visible_end = min(end, max_date)
                    is_active = visible_start <= d <= visible_end
                if is_active:
                    cell.fill = fill
                elif d == today:
                    cell.fill = today_fill

                if status_overlay:
                    labels = []
                    if testing_date == d:
                        labels.append("T")
                    if deploy_date == d:
                        labels.append("D")
                    if end and d == end and end <= today:
                        if not deploy_date or deploy_date > end:
                            if not testing_date or testing_date > end:
                                labels.append("!")
                            else:
                                labels.append("ϟ")
                    if labels:
                        black_font = InlineFont(rFont='Arial Unicode MS', b=True, sz=10, color=Color(rgb='FF000000'))
                        danger_font = InlineFont(rFont='Arial', b=True, sz=14, color=Color(rgb='FF000000'))
                        blocks = []
                        for idx, label in enumerate(labels):
                            if idx > 0:
                                blocks.append(TextBlock(black_font, " "))
                            if label == "!":
                                blocks.append(TextBlock(danger_font, label))
                            else:
                                blocks.append(TextBlock(black_font, label))
                        cell.value = CellRichText(blocks)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Форматирование строк
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                if cell.column == 2:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "G2"

        # ── Лист Задачи ──
        ws_tasks = wb.create_sheet("Задачи")
        ws_tasks.append([
            "#", "Проект", "Название", "Описание", "Статус", "Приоритет",
            "Ответственный", "Начало", "Дедлайн", "Тестирование", "Деплой", "Теги", "Список", "Создано", "Закрыто"
        ])
        for row in ws_tasks[1]:
            row.fill = header_fill
            row.font = Font(bold=True)
            row.alignment = Alignment(horizontal="center", vertical="center")

        for t in tasks:
            testing_date_str = ""
            deploy_date_str = ""
            if status_overlay:
                mapping = status_overlay_map.get(t.project_id, {})
                testing_date = get_first_status_date(t.id, mapping.get("testing"))
                deploy_date = get_first_status_date(t.id, mapping.get("deploy"))
                if testing_date:
                    testing_date_str = testing_date.strftime("%d.%m.%Y")
                if deploy_date:
                    deploy_date_str = deploy_date.strftime("%d.%m.%Y")

            ws_tasks.append([
                t.id,
                t.project.name if t.project else "",
                (t.title or "").strip() or (t.description or "").strip() or "—",
                t.description or "",
                t.status.name if t.status else "",
                t.priority or "",
                _format_task_assignees(t, assignees),
                t.start_date.strftime("%d.%m.%Y %H:%M") if t.start_date else "",
                t.due_date.strftime("%d.%m.%Y %H:%M") if t.due_date else "",
                testing_date_str,
                deploy_date_str,
                t.tags or "",
                t.list_name or "",
                t.created_at.strftime("%d.%m.%Y %H:%M") if t.created_at else "",
                "Да" if t.is_closed else "Нет",
            ])

        for col in ws_tasks.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws_tasks.column_dimensions[col_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"gantt_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.exception("Gantt XLSX export failed")
        raise HTTPException(status_code=500, detail=f"Ошибка формирования XLSX: {e}")
